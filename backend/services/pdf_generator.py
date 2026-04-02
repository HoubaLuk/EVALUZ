import os
import io
import json
from datetime import datetime
from fpdf import FPDF
from sqlalchemy.orm import Session
from models.db_models import StudentEvaluation, Lecturer, EvaluationCriteria, Criterion, ClassRoom, ClassAnalysis
from utils.sorting import sort_evaluations_by_surname


# Paths (Robust calculation for both local and Docker)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# BASE_DIR is now 'backend' folder
PUBLIC_DIR = os.path.join(os.path.dirname(BASE_DIR), "public")
FONTS_DIR = os.path.join(BASE_DIR, "static", "fonts")
LOGO_PATH = os.path.join(PUBLIC_DIR, "ÚPVSP bez okrajů.jpg")

def _parse_json_field(raw) -> dict:
    """Bezpečně parsuje JSON pole z DB (TEXT, může být double-encoded)."""
    if isinstance(raw, dict):
        return raw
    if not raw:
        return {}
    try:
        result = json.loads(raw)
        if isinstance(result, str):
            result = json.loads(result)
        return result if isinstance(result, dict) else {}
    except Exception:
        return {}


# Pomocná funkce pro převod plného názvu hodnosti na zkratku
def get_rank_abbr(full_rank: str) -> str:
    """
    Převede plný název hodnosti (např. 'vrchní komisař') na oficiální zkratku ('kpt.').
    Pokud hodnost není v číselníku, vrátí původní řetězec.
    """
    mapping = {
        "referent": "rtn.",
        "vrchní referent": "stržm.",
        "asistent": "nstržm.",
        "vrchní asistent": "pprap.",
        "inspektor": "prap.", 
        "komisař": "por.",
        "vrchní komisař": "kpt."
    }
    return mapping.get(full_rank.lower().strip(), full_rank)

class PDFReport(FPDF):
    """
    Rozšířená třída FPDF pro definici společného záhlaví a zápatí všech EVALUZ dokumentů.
    """
    def __init__(self, title="Globální analýza třídy - EVALUZ", subtitle="", unit="mm", format="A4"):
        super().__init__(unit=unit, format=format)
        self.report_title = title
        self.report_subtitle = subtitle

    def header(self):
        """Definuje záhlaví dokumentu (logo ÚPVSP a název systému). Zobrazuje se pouze na první straně."""
        if self.page_no() == 1:
            logo_path = os.path.join(BASE_DIR, "static", "logo_upvsp.png")
            if os.path.exists(logo_path):
                self.image(logo_path, x=10, y=8, w=20)
            
            self.set_font("DejaVu", "B", 10)
            self.set_xy(32, 10)
            self.cell(0, 6, "Útvar policejního vzdělávání a služební přípravy", border=0, ln=1)
            
            self.set_font("DejaVu", "", 10)
            self.set_xy(32, 16)
            self.cell(0, 6, self.report_title, border=0, ln=1)

            if getattr(self, "report_subtitle", ""):
                self.set_font("DejaVu", "", 8)
                self.set_text_color(100, 116, 139) # slate
                self.set_xy(32, 21)
                self.cell(0, 5, self.report_subtitle, border=0, ln=1)
                self.set_text_color(0, 0, 0)

            self.ln(10)

    def footer(self):
        """Definuje oficiální patičku s číslováním stránek a informací o vývoji."""
        self.set_y(-15)
        self.set_font("DejaVu", "", 8)
        self.set_text_color(100, 116, 139) # #64748b - břidlicově šedá
        self.cell(0, 5, "Generováno systémem EVALUZ vyvinutým Útvarem policejního vzdělávání a služební přípravy.", align="C", ln=1)
        self.cell(0, 5, f"Strana {self.page_no()}/{{nb}}", align="C", ln=1)
        self.set_text_color(0, 0, 0) # reset barvy na černou

def generate_student_pdf(evaluation: StudentEvaluation, lecturer: Lecturer, db: Session = None, scenario_id: str = "Nespecifikováno") -> bytes:
    """
    Generuje individuální hodnotící list studenta ve formátu PDF.
    
    Args:
        evaluation: Objekt evaluace z databáze obsahující JSON s výsledky.
        lecturer: Objekt lektora pro automatické doplnění podpisové doložky.
        db: Session databáze pro dohledání počtu bodů v kritériích.
        
    Returns:
        bytes: Binární data vygenerovaného PDF dokumentu.
    """
    def safe_text(txt):
        """Zajišťuje bezpečný převod textu pro PDF stroj."""
        if not txt: return ""
        return str(txt)

    # json_result je uložen jako TEXT — potřebujeme json.loads (ochrana před double-encoding)
    _raw = evaluation.json_result
    if isinstance(_raw, str):
        try:
            _raw = json.loads(_raw)
            if isinstance(_raw, str):
                _raw = json.loads(_raw)
        except Exception:
            _raw = {}
    data = _raw if isinstance(_raw, dict) else {}
    if not data:
        data = {"celkove_skore": 0, "zpetna_vazba": "Chyba formátu dat v databázi.", "vysledky": []}
    
    pdf = PDFReport(
        title="Hodnotící list vypracovaného úředního záznamu",
        subtitle=scenario_id, 
        unit="mm", format="A4"
    )
    
    # Cesty k TTF fontům podporujícím českou diakritiku
    font_regular = os.path.join(FONTS_DIR, "DejaVuSans.ttf")
    font_bold = os.path.join(FONTS_DIR, "DejaVuSans-Bold.ttf")
    font_oblique = os.path.join(FONTS_DIR, "DejaVuSans-Oblique.ttf")
    
    for fp in [font_regular, font_bold, font_oblique]:
        if not os.path.exists(fp):
            raise Exception(f"Font nenalezen na: {fp}")
            
    # Registrace fontů do FPDF instance
    pdf.add_font("DejaVu", "", font_regular)
    pdf.add_font("DejaVu", "B", font_bold)
    pdf.add_font("DejaVu", "I", font_oblique)
    
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # --- Sekce HLAVIČKA (Identifikace studenta) ---
    # Jméno: student_identity → cleaned_name → student_name (= název souboru, jen fallback)
    _student_display = ""
    if evaluation.student_identity:
        try:
            _id = evaluation.student_identity
            if isinstance(_id, str):
                _id = json.loads(_id)
            if isinstance(_id, str):
                _id = json.loads(_id)
            if isinstance(_id, dict) and _id.get("prijmeni"):
                _parts = [_id.get("hodnost", ""), _id.get("prijmeni", "").upper(), _id.get("jmeno", "")]
                _student_display = " ".join(p for p in _parts if p)
        except Exception:
            pass
    if not _student_display:
        _student_display = evaluation.cleaned_name or evaluation.student_name or ""

    pdf.set_font("DejaVu", "B", 12)
    pdf.cell(35, 8, "Student:", 0, 0)
    pdf.set_font("DejaVu", "", 12)
    pdf.cell(0, 8, safe_text(_student_display), 0, 1)
    
    pdf.ln(5)
    
    pdf.set_font("DejaVu", "B", 12)
    pdf.cell(0, 8, "Datum hodnocení:", 0, 1)
    pdf.set_font("DejaVu", "", 12)
    pdf.cell(0, 8, safe_text(datetime.now().strftime("%d. %m. %Y")), 0, 1)
    
    pdf.ln(2)
    
    # --- Sekce SKÓRE ---
    pdf.set_font("DejaVu", "B", 12)
    pdf.cell(35, 8, "Skóre:", 0, 0)
    pdf.set_font("DejaVu", "B", 14)
    pdf.set_text_color(0, 40, 85) # EVALUZ tmavě modrá
    
    # Výpočet max. bodů ze šablony kritérií v DB
    max_points = 25
    if db:
        crit_record = db.query(EvaluationCriteria).filter(
            EvaluationCriteria.scenario_name == scenario_id,
            
        ).first()
        if crit_record:
            criteria = db.query(Criterion).filter(Criterion.evaluation_criteria_id == crit_record.id).all()
            if criteria:
                max_points = sum(c.body for c in criteria if c.body)
                
    pdf.cell(0, 8, f"{data.get('celkove_skore', 0)} z {max_points} možných bodů", 0, 1)
    pdf.set_text_color(0, 0, 0) # reset
    
    pdf.ln(5)
    
    # --- Sekce TABULKA KRITÉRIÍ ---
    pdf.set_font("DejaVu", "B", 11)
    pdf.set_fill_color(0, 40, 85)
    pdf.set_text_color(255, 255, 255)
    
    # Definice šířek sloupců pro A4
    col_kriterium = 51
    col_splneno = 14
    col_body = 10
    col_oduv = 115
    
    pdf.cell(col_kriterium, 10, safe_text("Kritérium"), border=1, fill=True)
    pdf.cell(col_splneno, 10, safe_text("Splněno"), border=1, align="C", fill=True)
    pdf.cell(col_body, 10, "Body", border=1, align="C", fill=True)
    pdf.cell(col_oduv, 10, safe_text("Zdůvodnění AI"), border=1, fill=True)
    pdf.ln()

    # Vykreslení těla tabulky
    pdf.set_font("DejaVu", "", 10)
    pdf.set_text_color(0, 0, 0)
    
    for item in data.get("vysledky", []):
        nazev = safe_text(item.get("nazev", ""))
            
        splneno = "ANO" if item.get("splneno", False) else "NE"
        body = str(item.get("body", 0))
        oduvodneni = safe_text(item.get("oduvodneni", ""))
        citace = safe_text(item.get("citace", ""))
        
        # Sestavení finálního textu zdůvodnění s citací
        oduvodneni_text = oduvodneni.strip()
        if citace and citace.lower() not in ["", "chybí.", "chybí", "none", "(chybí)", "nenalezeno"]:
            oduvodneni_text += f"\n**Citace z textu:** {citace.strip()}"
            
        # Dynamický výpočet výšky řádku podle textu
        def get_lines_count(text, cell_width, is_markdown=False):
            if not text: return 1
            lines = pdf.multi_cell(cell_width, 6, text, dry_run=True, output="LINES", markdown=is_markdown)
            return len(lines)
            
        lines_oduv = get_lines_count(oduvodneni_text, col_oduv - 4, is_markdown=True)
        lines_nazev = get_lines_count(nazev, col_kriterium - 4, is_markdown=False)
        
        max_lines = max(lines_oduv, lines_nazev)
        row_h = (max_lines * 6) + 4
        if row_h < 12: row_h = 12
        
        x_before = pdf.get_x()
        y_before = pdf.get_y()
        
        # Automatické zalomení stránky uvnitř tabulky
        if y_before + row_h > 270:
            pdf.add_page()
            pdf.set_y(20)
            y_before = pdf.get_y()
            x_before = pdf.get_x()
            pdf.line(x_before, y_before, x_before + col_kriterium + col_splneno + col_body + col_oduv, y_before)
            
        # Vlastní vykreslení ohraničení pro sjednocenou výšku řádku
        pdf.line(x_before, y_before, x_before, y_before + row_h)
        pdf.line(x_before + col_kriterium, y_before, x_before + col_kriterium, y_before + row_h)
        pdf.line(x_before + col_kriterium + col_splneno, y_before, x_before + col_kriterium + col_splneno, y_before + row_h)
        pdf.line(x_before + col_kriterium + col_splneno + col_body, y_before, x_before + col_kriterium + col_splneno + col_body, y_before + row_h)
        pdf.line(x_before + col_kriterium + col_splneno + col_body + col_oduv, y_before, x_before + col_kriterium + col_splneno + col_body + col_oduv, y_before + row_h)
        
        # Vložení textu do buněk
        pdf.set_xy(x_before + 2, y_before + 2)
        pdf.multi_cell(col_kriterium - 4, 6, nazev, border=0)
        
        pdf.set_xy(x_before + col_kriterium, y_before)
        if splneno == "ANO": pdf.set_text_color(0, 150, 0) # zelená
        else: pdf.set_text_color(200, 0, 0) # červená
        pdf.cell(col_splneno, row_h, splneno, border=0, align="C")
        pdf.set_text_color(0, 0, 0)
        
        pdf.set_xy(x_before + col_kriterium + col_splneno, y_before)
        pdf.cell(col_body, row_h, body, border=0, align="C")
        
        pdf.set_xy(x_before + col_kriterium + col_splneno + col_body + 2, y_before + 2)
        pdf.multi_cell(col_oduv - 4, 6, oduvodneni_text, border=0, markdown=True)
        
        pdf.set_y(y_before + row_h)
        pdf.line(x_before, pdf.get_y(), x_before + col_kriterium + col_splneno + col_body + col_oduv, pdf.get_y())

    # --- Sekce ZPĚTNÁ VAZBA ---
    pdf.ln(8)
    pdf.set_font("DejaVu", "B", 12)
    pdf.cell(0, 8, safe_text("Celková zpětná vazba a doporučení:"), ln=1)
    pdf.set_font("DejaVu", "", 11)
    pdf.multi_cell(0, 6, safe_text(data.get("zpetna_vazba", "")), markdown=True)
    
    # --- Sekce PODPISOVÁ DOLOŽKA (ETŘ Kompatibilita) ---
    pdf.ln(12)
    pdf.set_font("DejaVu", "", 11)
    
    # Formát podpisové doložky: "Vyučující: kpt. Mgr. Lukáš Hřibňák, Ph.D." — bez funkcni_zarazeni
    _rank = (lecturer.rank_shortcut or "").strip()
    _tb = (lecturer.title_before or "").strip()
    _ta = (lecturer.title_after or "").strip()
    _fn = f"{lecturer.first_name} {lecturer.last_name}".strip()
    name_str = " ".join(filter(None, [_rank, _tb, _fn]))
    if _ta:
        name_str = f"{name_str}, {_ta}"

    pdf.cell(0, 8, safe_text(f"Vyučující: {name_str}"), ln=1)
    
    if lecturer.school_location:
        pdf.cell(0, 8, safe_text(lecturer.school_location), ln=1)
    
    # Právní a technická doložka v zápatí
    pdf.set_y(-25)
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("DejaVu", "I", 8)
    pdf.multi_cell(0, 5, safe_text("Zpracování dat probíhá v zabezpečeném vnitřním prostředí PČR v síti HERMES."))
    
    return pdf.output(dest="S")


def generate_class_excel(class_id: int, db: Session, current_user, scenario_id: str = None, class_name: str = "", scenario_display_name: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    cached_analysis = None
    if scenario_id:
        cached_analysis = db.query(ClassAnalysis).filter(ClassAnalysis.scenario_id == scenario_id).first()

    from api.auth import apply_data_isolation
    query = db.query(StudentEvaluation).filter(StudentEvaluation.class_id == class_id)
    evaluations = apply_data_isolation(query, StudentEvaluation, current_user, db).all()

    valid_evaluations = []
    for e in evaluations:
        if scenario_id and e.scenario_name != scenario_id:
            continue
        data = _parse_json_field(e.json_result)
        if data and data.get("vysledky"):
            valid_evaluations.append(e)
            
    evaluations = sort_evaluations_by_surname(valid_evaluations)

    # Získání definic kritérií z DB

    db_criteria = []
    if scenario_id:
        criteria_record = db.query(EvaluationCriteria).filter(EvaluationCriteria.scenario_name == scenario_id).first()
        if criteria_record:
            from models.db_models import Criterion
            db_criteria = db.query(Criterion).filter(Criterion.evaluation_criteria_id == criteria_record.id).all()
            
    criteria_names = [c.nazev for c in db_criteria]
    max_pts = sum([c.body for c in db_criteria]) if db_criteria else 0
    
    if not criteria_names:
        for eval_record in evaluations:
            try:
                data = eval_record.json_result or {}
                if "vysledky" in data and isinstance(data["vysledky"], list):
                    for item in data["vysledky"]:
                        if "nazev" in item and item["nazev"] not in criteria_names:
                            criteria_names.append(item["nazev"])
            except Exception:
                pass

    wb = Workbook()
    
    header_fill = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    footer_text = "Generováno systémem EVALUZ - Vyvinuto na ÚPVSP"

    # --- LIST 1: Souhrn ---
    ws_summary = wb.active
    ws_summary.title = "Souhrn"
    
    # Stylizované Záhlaví
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = "PROTOKOL O HODNOCENÍ TŘÍDY - EVALUZ"
    ws_summary['A1'].font = Font(size=14, bold=True, color="002855")
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    # Třída: query param má prioritu, fallback na DB
    _trida = class_name
    if not _trida:
        _cr = db.query(ClassRoom).filter(ClassRoom.id == class_id).first()
        _trida = _cr.name if _cr else "Neznámá"
    # Modelová situace: query param má prioritu, fallback na scenario_id
    _situace = scenario_display_name or scenario_id or "Neznámá"
    ws_summary.append(["Třída:", _trida])
    ws_summary.append(["Modelová situace:", _situace])
    ws_summary.append(["Datum exportu:", datetime.now().strftime("%d. %m. %Y %H:%M")])
    ws_summary.append([]) # Mezera
    
    start_row = ws_summary.max_row + 1
    ws_summary.append(["Statistika", "Hodnota"])
    for cell in ws_summary[start_row]:
        cell.fill = header_fill
        cell.font = header_font


    total_students = len(evaluations)
    total_score_sum = 0
    dist = {"0-50%": 0, "51-80%": 0, "81-100%": 0}

    for eval_record in evaluations:
        try:
            data = _parse_json_field(eval_record.json_result)
            score = data.get('celkove_skore', 0)
            total_score_sum += score
            if max_pts > 0:
                percent = (score / max_pts) * 100
                if percent < 50: dist["0-50%"] += 1
                elif percent <= 80: dist["51-80%"] += 1
                else: dist["81-100%"] += 1
        except Exception:
            pass

    avg_score = round(total_score_sum / total_students, 1) if total_students > 0 else 0
    ws_summary.append(["Celkový počet hodnocených studentů", total_students])
    ws_summary.append(["Max. možný počet bodů", max_pts])
    
    # Průměrné skóre jako číslo pro automatické zarovnání vpravo
    ws_summary.append(["Průměrné skóre (body)", avg_score])
    ws_summary.cell(row=ws_summary.max_row, column=2).number_format = '#,##0.0'
    
    ws_summary.append(["Úspěšnost (pod 50 %)", dist["0-50%"]])
    ws_summary.append(["Úspěšnost (51-80 %)", dist["51-80%"]])
    ws_summary.append(["Úspěšnost (81-100 %)", dist["81-100%"]])
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.append([])
    ws_summary.append([footer_text])

    # --- LIST 2: Výsledky ---
    ws_results = wb.create_sheet(title="Výsledky")
    headers = ["Jméno Studenta"] + [f"K{i+1}" for i in range(len(criteria_names))] + ["Celkové Skóre"]
    ws_results.append(headers)
    for cell in ws_results[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_results.freeze_panes = "B2"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    crit_totals = {name: {"passes": 0, "pts": 0} for name in criteria_names}
    
    for eval_record in evaluations:
        data = _parse_json_field(eval_record.json_result)

        # Priorita: cleaned_name -> student_name
        student_display_name = eval_record.cleaned_name if eval_record.cleaned_name else eval_record.student_name
        total_score = data.get('celkove_skore', 0)

        
        points = []
        parsed_results = data.get("vysledky", [])
        for critique_name in criteria_names:
            point_value = 0
            for item in parsed_results:
                if item.get("nazev") == critique_name:
                    point_value = item.get("body", 0)
                    if item.get("splneno"):
                        crit_totals[critique_name]["passes"] += 1
                    break
            points.append(point_value)
            
        row_data = [student_display_name] + points + [total_score]
        ws_results.append(row_data)

        
        last_row = ws_results.max_row
        for col_idx in range(2, len(points) + 2):
            val = ws_results.cell(row=last_row, column=col_idx).value
            if val == 0:
                ws_results.cell(row=last_row, column=col_idx).fill = red_fill
            else:
                ws_results.cell(row=last_row, column=col_idx).fill = green_fill

    for col in ws_results.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws_results.column_dimensions[column].width = (max_length + 2) * 1.2
    
    # --- LIST 3: Analýza ---
    ws_analysis = wb.create_sheet(title="Analýza")
    ws_analysis.append(["Kód", "Znění Kritéria", "Úspěšnost (%)"])
    for cell in ws_analysis[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for idx, crit_name in enumerate(criteria_names):
        passes = crit_totals[crit_name]["passes"]
        pct = round((passes / total_students) * 100) if total_students > 0 else 0
        ws_analysis.append([f"K{idx+1}", crit_name, pct])

    ws_analysis.column_dimensions['A'].width = 10
    ws_analysis.column_dimensions['B'].width = 80
    ws_analysis.column_dimensions['C'].width = 15
    
    # Pridani Grafu
    if total_students > 0 and criteria_names:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Úspěšnost kritérií"
        chart.y_axis.title = "Úspěšnost (%)"
        chart.x_axis.title = "Kritéria"
        
        data_ref = Reference(ws_analysis, min_col=3, min_row=1, max_row=len(criteria_names)+1)
        cats = Reference(ws_analysis, min_col=1, min_row=2, max_row=len(criteria_names)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_analysis.add_chart(chart, "E2")
    
    # --- LIST 4: Metodika (Odstraníme nebo ponecháme jen pokud je obsah) ---
    if cached_analysis and cached_analysis.content_json:
        ws_methodology = wb.create_sheet(title="Metodika")
        ws_methodology.append(["Pedagogické shrnutí od AI Asistenta"])
        for cell in ws_methodology[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        try:
            analysis_data = _parse_json_field(cached_analysis.content_json)
            ai_insight = analysis_data.get("ai_insight", "")
            if not ai_insight:
                raise ValueError("Chybí ai_insight")
            blocks = ai_insight.split("###")
            for block in blocks:
                block = block.strip()
                if not block: continue
                lines = block.split("\n")
                if lines:
                    ws_methodology.append([lines[0].replace("**", "").strip()])
                    ws_methodology.cell(row=ws_methodology.max_row, column=1).font = Font(bold=True)
                    content = "\n".join(lines[1:]).strip()
                    ws_methodology.append([content])
                    ws_methodology.append([])
        except Exception:
            ws_methodology.append(["Shrnutí není k dispozici — nejprve vygenerujte analýzu ve frontendu."])

        ws_methodology.column_dimensions['A'].width = 100
        for cell in ws_methodology['A']:
            cell.alignment = Alignment(wrap_text=True)


    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_class_report_pdf(
    analysis_data: dict,
    scenario_id: str,
    lecturer: Lecturer,
    scenario_display_name: str = "",
    class_name: str = "",
    criteria_descriptions: dict = None
) -> bytes:
    """
    Generuje komplexní PDF report s globální analýzou celé třídy.
    Obsahuje KPI karty, tabulku úspěšnosti kritérií s barevným kódováním
    a hloubkový AI pedagogický vhled.
    """
    try:
        final_scenario_id = analysis_data.get("scenario_id") or scenario_id
        scenario_name = f"Zvolený scénář (ID: {final_scenario_id})"

        pdf = PDFReport(
            title="Protokol o hodnocení studijní skupiny - EVALUZ",
            unit="mm", format="A4"
        )

        # Registrace fontů
        font_regular = os.path.join(FONTS_DIR, "DejaVuSans.ttf")
        font_bold = os.path.join(FONTS_DIR, "DejaVuSans-Bold.ttf")
        font_oblique = os.path.join(FONTS_DIR, "DejaVuSans-Oblique.ttf")
        
        if os.path.exists(font_regular):
            pdf.add_font('DejaVu', '', font_regular)
            pdf.add_font('DejaVu', 'B', font_bold)
            pdf.add_font('DejaVu', 'I', font_oblique)
        else:
            # V Dockeru / PC bez DejaVu fallback na standardní fonty (bez UTF-8 podpory pro CZ, ale nespadne to)
            pdf.set_font("Helvetica", size=10)
            print(f">>> [PDF WARNING] Fonty nenalezeny v {FONTS_DIR}, pouzivam Helvetica fallback.")
            
        pdf.add_page()
        
        # --- Záhlaví Analýzy ---
        pdf.set_draw_color(30, 41, 59)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("DejaVu", "B", 18)
        pdf.cell(0, 16, "  Protokol o hodnocení studijní skupiny", border=0, ln=1, align="L", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("DejaVu", "", 10)
        pdf.set_draw_color(0, 0, 0)
        pdf.ln(6)
        
        # Hlavička informací — 5 polí
        pdf.set_fill_color(240, 245, 250)
        import datetime
        now_str = datetime.datetime.now().strftime("%d. %m. %Y %H:%M")

        # Formát: "kpt. Mgr. Lukáš Hřibňák, Ph.D." — bez funkcni_zarazeni
        rank_part = (lecturer.rank_shortcut or "").strip()
        title_before = (lecturer.title_before or "").strip()
        title_after = (lecturer.title_after or "").strip()
        full_name = f"{lecturer.first_name} {lecturer.last_name}".strip()
        name_str = " ".join(filter(None, [rank_part, title_before, full_name]))
        if title_after:
            name_str = f"{name_str}, {title_after}"

        label_w = 50  # šířka popisku

        pdf.cell(label_w, 8, "Vzdělávací zařízení:", border=1, fill=True)
        pdf.cell(0, 8, lecturer.school_location or "—", border=1, ln=1)

        pdf.cell(label_w, 8, "Studijní skupina:", border=1, fill=True)
        pdf.cell(0, 8, class_name or "—", border=1, ln=1)

        pdf.cell(label_w, 8, "Modelová situace:", border=1, fill=True)
        pdf.cell(0, 8, scenario_display_name or scenario_name, border=1, ln=1)

        pdf.cell(label_w, 8, "Vyučující:", border=1, fill=True)
        pdf.cell(0, 8, name_str or "—", border=1, ln=1)

        pdf.cell(label_w, 8, "Datum exportu:", border=1, fill=True)
        pdf.cell(0, 8, now_str, border=1, ln=1)
    
        # KPI Card
        pdf.ln(6)
        avg_score = analysis_data.get("average_score", 0)
        max_score = analysis_data.get("max_score", 0)
        pdf.set_font("DejaVu", "B", 16)
        pdf.set_fill_color(241, 245, 249) # slate-100
        pdf.set_text_color(15, 23, 42)    # slate-900
        pdf.set_draw_color(203, 213, 225) # slate-300
        pdf.cell(0, 18, f"  Průměrné skóre třídy:  {avg_score} / {max_score} bodů", border=1, ln=1, align="C", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_draw_color(0, 0, 0)
        
        # Shrnutí tabulky K1-K25
        pdf.ln(10)
        pdf.set_font("DejaVu", "B", 12)
        pdf.cell(0, 8, "Úspěšnost plnění kritérií", ln=1)

        # Pomocná funkce pro výpočet výšky řádku
        def calc_row_height(pdf_ref, text, col_w, line_h=5):
            """Vypočítá potřebnou výšku řádku dle délky textu."""
            words = text.split()
            current_w = 0
            lines = 1
            for word in words:
                w = pdf_ref.get_string_width(word + ' ')
                if current_w + w > col_w - 2:
                    lines += 1
                    current_w = w
                else:
                    current_w += w
            return max(line_h, lines * line_h)

        total_students = sum(analysis_data.get("score_distribution", {}).values()) or 1

        pdf.set_font("DejaVu", "B", 9)
        pdf.set_draw_color(100, 116, 139)  # slate-500
        pdf.set_fill_color(241, 245, 249)  # slate-100
        pdf.cell(22, 8, "Kritérium č.", border=1, fill=True, align="C")
        pdf.cell(130, 8, "Definice kritéria", border=1, fill=True)
        pdf.cell(0, 8, "Splnilo", border=1, ln=1, align="C", fill=True)

        pdf.set_font("DejaVu", "", 8)
        pdf.set_draw_color(203, 213, 225)  # slate-300 for rows
        stats = analysis_data.get("stats", [])
        remaining_w = pdf.w - pdf.l_margin - pdf.r_margin - 22 - 130

        for i, stat in enumerate(stats):
            success_rate = stat.get('success_rate', 0)
            passes = round(success_rate / 100 * total_students)
            splnilo = f"{passes} ({success_rate} %)"

            full_name = stat.get('full_name', stat.get('name', ''))
            description = full_name  # krátký název kritéria (c.nazev)

            row_h = calc_row_height(pdf, description, 130)
            x0, y0 = pdf.get_x(), pdf.get_y()

            # Kontrola konce stránky
            if y0 + row_h > pdf.h - pdf.b_margin - 10:
                pdf.add_page()
                x0, y0 = pdf.get_x(), pdf.get_y()

            # Policejní 5-stupňová škála (barvy sloupce "Splnilo")
            if success_rate >= 80:
                fill_rgb = (220, 252, 231); text_rgb = (22, 101, 52)    # 1 — zelená
            elif success_rate >= 60:
                fill_rgb = (187, 247, 208); text_rgb = (21, 128, 61)    # 2 — světle zelená
            elif success_rate >= 40:
                fill_rgb = (254, 243, 199); text_rgb = (146, 64, 14)    # 3 — amber
            elif success_rate >= 20:
                fill_rgb = (254, 215, 170); text_rgb = (154, 52, 18)    # 4 — oranžová
            else:
                fill_rgb = (254, 226, 226); text_rgb = (153, 27, 27)    # 5 — červená

            # Sloupec 1: Kritérium č.
            pdf.set_xy(x0, y0)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(22, row_h, f"K{i+1}", border=1, align="C")

            # Sloupec 2: Definice kritéria (multi_cell – zalamuje text)
            pdf.set_xy(x0 + 22, y0)
            pdf.multi_cell(130, 5, description, border=1)

            # Sloupec 3: Splnilo — umístit na správnou pozici se správnou výškou
            pdf.set_xy(x0 + 22 + 130, y0)
            pdf.set_fill_color(*fill_rgb)
            pdf.set_text_color(*text_rgb)
            pdf.set_font("DejaVu", "B", 8)
            pdf.cell(remaining_w, row_h, splnilo, border=1, ln=0, align="C", fill=True)

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("DejaVu", "", 8)
            pdf.set_fill_color(255, 255, 255)
            pdf.set_y(y0 + row_h)

        # AI Pedagogické shrnutí — bez velké mezery, rovnou pokračuje
        pdf.ln(4)
        pdf.set_font("DejaVu", "B", 14)
        pdf.set_text_color(0, 40, 85)
        pdf.cell(0, 10, "PEDAGOGICKÉ SHRNUTÍ (AI ASISTENT)", border=0, ln=1, align="L")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)
        
        pdf.set_font("DejaVu", "", 10)
        ai_insight = analysis_data.get("ai_insight", "")
        
        if ai_insight:
            # Split blocks
            blocks = ai_insight.split("###")
            for block in blocks:
                block = block.strip()
                if not block: continue
                
                lines = block.split("\n")
                if lines:
                    title = lines[0].replace("**", "").strip()
                    content = "\n".join(lines[1:]).strip()
                    
                    pdf.set_font("DejaVu", "B", 12)
                    pdf.set_text_color(30, 41, 59)
                    pdf.cell(0, 8, title, ln=1)
                    pdf.set_text_color(0, 0, 0)
                    
                    pdf.set_font("DejaVu", "", 10)
                    # Styled grey boxes for AI insight paragraphs
                    pdf.set_fill_color(248, 250, 252) # slate-50
                    pdf.set_draw_color(226, 232, 240) # slate-200
                    pdf.multi_cell(0, 6, "\n" + content + "\n", markdown=True, border=1, fill=True)
                    pdf.ln(6)

        return bytes(pdf.output())

    except Exception as e:
        import traceback
        print(">>> [PDF GEN ERROR] Chycena výjimka při generování třídního reportu:")
        traceback.print_exc()
        raise Exception(f"Failed to generate PDF: {str(e)}")

def generate_dashboard_excel(data: dict) -> bytes:
    """
    Generuje vícestránkový Excel report ze statistik využití (dashboard).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import datetime

    wb = Workbook()
    
    header_fill = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    footer_text = "Generováno systémem EVALUZ - Vyvinuto na ÚPVSP"

    # --- LIST 1: Celkové shrnutí ---
    ws_summary = wb.active
    ws_summary.title = "Základní přehled"
    
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = "STATISTIKY VYUŽITÍ EVALUZ"
    ws_summary['A1'].font = Font(size=14, bold=True, color="002855")
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    ws_summary.append(["Role:", str(data.get("role", "")).upper()])
    ws_summary.append(["Organizační článek:", data.get("org_unit", "")])
    ws_summary.append(["Datum exportu:", datetime.datetime.now().strftime("%d. %m. %Y %H:%M")])
    ws_summary.append([]) # Mezera
    
    start_row = ws_summary.max_row + 1
    ws_summary.append(["Metrika", "Hodnota"])
    for cell in ws_summary[start_row]:
        cell.fill = header_fill
        cell.font = header_font

    ws_summary.append(["Celkový počet hodnocených ÚZ", data.get("total_evaluations", 0)])
    by_lec = data.get("by_lecturer", [])
    ws_summary.append(["Počet aktivních vyučujících", len(by_lec)])
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.append([])
    ws_summary.append([footer_text])

    # --- LIST 2: Organizační články ---
    if data.get("role") == "superadmin":
        ws_org = wb.create_sheet(title="Organizační články")
        ws_org.append(["Název článku", "Počet ÚZ"])
        for cell in ws_org[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        for row in data.get("by_org_unit", []):
            ws_org.append([row.get("name"), row.get("count")])
            
        ws_org.column_dimensions['A'].width = 60
        ws_org.column_dimensions['B'].width = 15
        
        ws_org.append([])
        ws_org.append([footer_text])

    # --- LIST 3: Lektoři ---
    ws_lec = wb.create_sheet(title="Aktivita vyučujících")
    ws_lec.append(["Jméno vyučujícího", "Počet vyhodnocených ÚZ"])
    for cell in ws_lec[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for row in data.get("by_lecturer", []):
        ws_lec.append([row.get("name"), row.get("count")])
        
    ws_lec.column_dimensions['A'].width = 40
    ws_lec.column_dimensions['B'].width = 25
    
    ws_lec.append([])
    ws_lec.append([footer_text])

    # --- LIST 4: Časová osa ---
    ws_time = wb.create_sheet(title="Časová osa")
    ws_time.append(["Datum", "Počet vyhodnocených ÚZ"])
    for cell in ws_time[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for row in data.get("timeline", []):
        ws_time.append([row.get("date"), row.get("count")])
        
    ws_time.column_dimensions['A'].width = 20
    ws_time.column_dimensions['B'].width = 25
    
    ws_time.append([])
    ws_time.append([footer_text])

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
