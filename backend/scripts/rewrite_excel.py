import sys
import os

# Adds the backend path
backend_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(backend_path)

file_path = os.path.join(backend_path, "services", "pdf_generator.py")
with open(file_path, "r", encoding="utf-8") as f:
    content = f.read()

import re

# Use regex to find generate_student_pdf and replace hardcoded MS2
content = content.replace('EvaluationCriteria.scenario_name == "MS2",', 'EvaluationCriteria.scenario_name == scenario_id,')
content = content.replace('EvaluationCriteria.lecturer_id == lecturer.id', '')

# Remove max_score = analysis_data.get("max_score", 25)
content = content.replace('max_score = analysis_data.get("max_score", 25)', 'max_score = analysis_data.get("max_score", 0)')
# Remove [:25]
content = content.replace('for i, stat in enumerate(stats[:25]):', 'for i, stat in enumerate(stats):')


# Rewrite generate_class_excel
excel_code = """
def generate_class_excel(class_id: int, db: Session, lecturer_id: int, scenario_id: str = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    evaluations = db.query(StudentEvaluation).filter(
        StudentEvaluation.class_id == class_id,
        StudentEvaluation.lecturer_id == lecturer_id
    ).all()

    if scenario_id:
        evaluations = [e for e in evaluations if e.scenario_name == scenario_id]

    # Získání definic kritérií z DB
    db_criteria = []
    if scenario_id:
        criteria_record = db.query(EvaluationCriteria).filter(EvaluationCriteria.scenario_name == scenario_id).first()
        if criteria_record:
            from models.db_models import Criterion
            db_criteria = db.query(Criterion).filter(Criterion.evaluation_criteria_id == criteria_record.id).all()
            
    criteria_names = [c.nazev for c in db_criteria]
    max_pts = sum([c.body for c in db_criteria]) if db_criteria else 0
    
    if not criteria_names:
        for eval_record in evaluations:
            try:
                data = json.loads(eval_record.json_result)
                if "vysledky" in data and isinstance(data["vysledky"], list):
                    for item in data["vysledky"]:
                        if "nazev" in item and item["nazev"] not in criteria_names:
                            criteria_names.append(item["nazev"])
            except Exception:
                pass

    wb = Workbook()
    
    header_fill = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    footer_text = "Generováno systémem EVALUZ - Vyvinuto na ÚPVSP"

    # --- LIST 1: Souhrn ---
    ws_summary = wb.active
    ws_summary.title = "Souhrn"
    ws_summary.append(["Statistika", "Hodnota"])
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    total_students = len(evaluations)
    total_score_sum = 0
    dist = {"0-50%": 0, "51-80%": 0, "81-100%": 0}

    for eval_record in evaluations:
        try:
            data = json.loads(eval_record.json_result)
            score = data.get('celkove_skore', 0)
            total_score_sum += score
            if max_pts > 0:
                percent = (score / max_pts) * 100
                if percent < 50: dist["0-50%"] += 1
                elif percent <= 80: dist["51-80%"] += 1
                else: dist["81-100%"] += 1
        except Exception:
            pass

    avg_score = round(total_score_sum / total_students, 1) if total_students > 0 else 0
    ws_summary.append(["Celkový počet hodnocených studentů", total_students])
    ws_summary.append(["Max. možný počet bodů", max_pts])
    ws_summary.append(["Průměrné skóre", f"{avg_score} b."])
    ws_summary.append(["Úspěšnost (pod 50 %)", dist["0-50%"]])
    ws_summary.append(["Úspěšnost (51-80 %)", dist["51-80%"]])
    ws_summary.append(["Úspěšnost (81-100 %)", dist["81-100%"]])
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.append([])
    ws_summary.append([footer_text])

    # --- LIST 2: Výsledky ---
    ws_results = wb.create_sheet(title="Výsledky")
    headers = ["Jméno Studenta"] + [f"K{i+1}" for i in range(len(criteria_names))] + ["Celkové Skóre"]
    ws_results.append(headers)
    for cell in ws_results[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_results.freeze_panes = "B2"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    crit_totals = {name: {"passes": 0, "pts": 0} for name in criteria_names}
    
    for eval_record in evaluations:
        try:
            data = json.loads(eval_record.json_result)
        except Exception:
            data = {}
        student_name = eval_record.student_name.replace(',', '')
        total_score = data.get('celkove_skore', 0)
        
        points = []
        parsed_results = data.get("vysledky", [])
        for critique_name in criteria_names:
            point_value = 0
            for item in parsed_results:
                if item.get("nazev") == critique_name:
                    point_value = item.get("body", 0)
                    if item.get("splneno"):
                        crit_totals[critique_name]["passes"] += 1
                    break
            points.append(point_value)
            
        row_data = [student_name] + points + [total_score]
        ws_results.append(row_data)
        
        last_row = ws_results.max_row
        for col_idx in range(2, len(points) + 2):
            val = ws_results.cell(row=last_row, column=col_idx).value
            if val == 0:
                ws_results.cell(row=last_row, column=col_idx).fill = red_fill
            else:
                ws_results.cell(row=last_row, column=col_idx).fill = green_fill

    for col in ws_results.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws_results.column_dimensions[column].width = (max_length + 2) * 1.2
    
    ws_results.append([])
    ws_results.append([footer_text])

    # --- LIST 3: Analýza ---
    ws_analysis = wb.create_sheet(title="Analýza")
    ws_analysis.append(["Kód", "Znění Kritéria", "Úspěšnost (%)"])
    for cell in ws_analysis[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for idx, crit_name in enumerate(criteria_names):
        passes = crit_totals[crit_name]["passes"]
        pct = round((passes / total_students) * 100) if total_students > 0 else 0
        ws_analysis.append([f"K{idx+1}", crit_name, pct])

    ws_analysis.column_dimensions['A'].width = 10
    ws_analysis.column_dimensions['B'].width = 80
    ws_analysis.column_dimensions['C'].width = 15
    
    # Pridani Grafu
    if total_students > 0 and criteria_names:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Úspěšnost kritérií"
        chart.y_axis.title = "Úspěšnost (%)"
        chart.x_axis.title = "Kritéria"
        
        data_ref = Reference(ws_analysis, min_col=3, min_row=1, max_row=len(criteria_names)+1)
        cats = Reference(ws_analysis, min_col=1, min_row=2, max_row=len(criteria_names)+1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_analysis.add_chart(chart, "E2")
    
    ws_analysis.append([])
    ws_analysis.append([footer_text])

    # --- LIST 4: Metodika ---
    ws_methodology = wb.create_sheet(title="Metodika")
    ws_methodology.append(["Pedagogické shrnutí od AI Asistenta"])
    for cell in ws_methodology[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Try fetching AI insight from ClassAnalysis
    from models.db_models import ClassAnalysis
    cached_analysis = db.query(ClassAnalysis).filter(ClassAnalysis.scenario_id == scenario_id).first()
    if cached_analysis:
        try:
            analysis_data = json.loads(cached_analysis.content_json)
            ai_insight = analysis_data.get("ai_insight", "")
            blocks = ai_insight.split("###")
            for block in blocks:
                block = block.strip()
                if not block: continue
                lines = block.split("\\n")
                if lines:
                    ws_methodology.append([lines[0].replace("**", "").strip()])
                    ws_methodology.cell(row=ws_methodology.max_row, column=1).font = Font(bold=True)
                    content = "\\n".join(lines[1:]).strip()
                    ws_methodology.append([content])
                    ws_methodology.append([])
        except Exception:
            ws_methodology.append(["Nepodařilo se načíst zprávu."])
            
    ws_methodology.column_dimensions['A'].width = 100
    for cell in ws_methodology['A']:
        cell.alignment = Alignment(wrap_text=True)
        
    ws_methodology.append([])
    ws_methodology.append([footer_text])

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
"""

# Replace the old generate_class_excel
import re
pattern = re.compile(r'def generate_class_excel\(class_id:(.*?)\ndef generate_class_report_pdf\(', re.DOTALL)
content = pattern.sub(excel_code + "\n\ndef generate_class_report_pdf(", content)

with open(file_path, "w", encoding="utf-8") as f:
    f.write(content)
print("Rewrote pdf_generator.py to use dynamic queries and openpyxl formatting.")
